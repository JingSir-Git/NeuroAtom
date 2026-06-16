"""BASE-AAD-USTC Importer: USTC dichotic auditory attention decoding.

18 normal-hearing subjects, dichotic listening (one speaker per ear), 20 trials
of 120 s each, attention switched between successive trials. Recorded on a
Neuroscan SynAmps (Quik-Cap 64) at 1000 Hz.

Why a custom reader: the recordings are Neuroscan Curry 8 ``.cdt`` files whose
header ``NumSamples`` disagrees with the actual file size, so ``mne.io.read_raw_curry``
crashes on them. We read the raw float32 payload directly (sample-multiplexed,
already in µV per the ``.dpa`` ``DataUnit``) and take channel labels + sampling
rate from the ASCII ``.dpa`` sidecar.

Trial segmentation comes from the Trigger channel: it holds an inter-trial
baseline code (65280) and steps to ``65280 + trial_number`` for the duration of
each trial. Per-trial attended ear (L/R) is read from ``group1.xlsx``.

Data layout::

    BASE-AAD-USTC/
        group1.xlsx                 # Subject / Trial / attended (L|R)
        Audio/                      # wav stimuli (not imported)
        EEG/
            s1.cdt  s1.cdt.dpa  s1.cdt.ceo ...
            s2.cdt  ...

Each trial → one TRIAL atom (64 EEG + HEO/VEO/EKG/EMG; Trigger excluded) with a
static ``attended_ear`` label. A handful of trials flagged by the dataset as
having poor behavioural answers are imported but annotated ``excluded=true``.
"""

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np

from neuroatom.core.annotation import CategoricalAnnotation, NumericAnnotation, TextAnnotation
from neuroatom.core.atom import Atom, TemporalInfo
from neuroatom.core.channel import ChannelInfo
from neuroatom.core.dataset_meta import DatasetMeta
from neuroatom.core.enums import AtomType, ChannelStatus, ChannelType
from neuroatom.core.run import RunMeta
from neuroatom.core.signal_ref import SignalRef
from neuroatom.core.subject import SubjectMeta
from neuroatom.importers.base import BaseImporter, ImportResult, TaskConfig
from neuroatom.importers.registry import register_importer
from neuroatom.utils.channel_names import standardize_channel_name
from neuroatom.utils.hashing import compute_atom_id
from neuroatom.utils.unit_convert import convert_to_storage_unit
from neuroatom.utils.validation import validate_signal

logger = logging.getLogger(__name__)

# Dataset-flagged trials with poor behavioural answers (subject → trial numbers).
_POOR_TRIALS = {"13": {16}, "14": {13}, "16": {4}, "17": {18}}


def _ch_type(name: str) -> ChannelType:
    u = name.upper()
    if u == "TRIGGER":
        return ChannelType.STIM
    if u in ("HEO", "VEO", "HEOG", "VEOG"):
        return ChannelType.EOG
    if u in ("M1", "M2"):
        return ChannelType.REF
    if u in ("EKG", "ECG", "EMG"):
        return ChannelType.OTHER
    return ChannelType.EEG


def _parse_dpa(dpa_path: Path) -> Dict[str, Any]:
    """Parse a Curry ``.dpa`` ASCII sidecar for params + channel labels."""
    text = dpa_path.read_text(encoding="latin-1", errors="replace")

    def _kv(key: str, default: Optional[str] = None) -> Optional[str]:
        m = re.search(rf"^\s*{key}\s*=\s*(.+?)\s*$", text, re.MULTILINE)
        return m.group(1).strip() if m else default

    def _list(tag: str) -> List[str]:
        m = re.search(rf"{tag} START_LIST.*?\n(.*?)\n\s*{tag} END_LIST", text, re.DOTALL)
        if not m:
            return []
        out = []
        for line in m.group(1).splitlines():
            line = line.split("#")[0].strip()
            if line:
                out.append(line)
        return out

    labels = _list("LABELS") + _list("LABELS_OTHERS")
    return {
        "n_channels": int(_kv("NumChannels", "0") or 0),
        "sfreq": float(_kv("SampleFreqHz", "1000") or 1000),
        "unit": (_kv("DataUnit", "uV") or "uV"),
        "labels": labels,
    }


def _read_cdt(cdt_path: Path, n_channels: int) -> np.ndarray:
    """Memory-map a ``.cdt`` payload as (n_samples, n_channels) float32 (µV).

    Uses the actual file size (the header's NumSamples is unreliable here).
    Sample-multiplexed order → row = sample, col = channel.
    """
    n_floats = cdt_path.stat().st_size // 4
    n_samples = n_floats // n_channels
    if n_samples * n_channels != n_floats:
        logger.warning(
            "%s: %d floats not divisible by %d channels (%d remainder)",
            cdt_path.name, n_floats, n_channels, n_floats % n_channels,
        )
    return np.memmap(
        cdt_path, dtype="<f4", mode="r", shape=(n_samples, n_channels),
    )


def _find_trial_onsets(trigger: np.ndarray) -> Dict[int, int]:
    """Map trial_number → onset_sample from the Trigger channel.

    The most frequent value is the inter-trial baseline; every other value is a
    brief pulse to ``baseline + trial_number`` marking that trial's onset (the
    code is *not* held for the trial duration). The trial window is taken as a
    fixed length forward from the onset by the caller.
    """
    trig = np.asarray(trigger).round().astype(np.int64)
    vals, counts = np.unique(trig, return_counts=True)
    baseline = int(vals[int(np.argmax(counts))])

    onsets: Dict[int, int] = {}
    for v in vals:
        v = int(v)
        if v == baseline:
            continue
        where = np.where(trig == v)[0]
        if where.size:
            onsets[v - baseline] = int(where[0])
    return onsets


def _parse_group_xlsx(xlsx_path: Path) -> Dict[str, Dict[int, Dict[str, str]]]:
    """Parse group1.xlsx → {subject: {trial_num: {attended, wav}}}."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    out: Dict[str, Dict[int, Dict[str, str]]] = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        current: Optional[str] = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row or all(c is None for c in row[:3]):
                continue
            if row[0] is not None:
                current = str(row[0]).strip()
                out.setdefault(current, {})
            if current is None or row[1] is None:
                continue
            m = re.match(r"\s*(\d+)", str(row[1]))
            if not m:
                continue
            trial_num = int(m.group(1))
            wav = str(row[1]).split("、")[-1].strip() if "、" in str(row[1]) else str(row[1]).strip()
            attended = str(row[2]).strip().upper() if row[2] is not None else ""
            out[current][trial_num] = {"attended": attended, "wav": wav}
    return out


class BASEAADUSTCImporter(BaseImporter):
    """Importer for the BASE-AAD-USTC dichotic AAD dataset (custom Curry reader)."""

    @staticmethod
    def detect(path: Path) -> bool:
        path = Path(path)
        if path.is_file():
            return path.suffix.lower() == ".cdt" and (path.parent.name == "EEG")
        if path.is_dir():
            if (path / "EEG").is_dir() and any((path / "EEG").glob("*.cdt")):
                return (path / "group1.xlsx").exists()
            if path.name == "EEG" and any(path.glob("*.cdt")):
                return True
        return False

    def load_raw(self, path: Path) -> Tuple[Any, Dict[str, Any]]:
        raise NotImplementedError("BASE-AAD-USTC uses import_subject().")

    def extract_channel_infos(self, raw: Any) -> List[ChannelInfo]:
        return []

    def extract_events(self, raw: Any) -> Optional[np.ndarray]:
        return None

    # ------------------------------------------------------------------

    def _build_channel_infos(
        self, labels: List[str], sfreq: float,
    ) -> Tuple[List[ChannelInfo], List[int]]:
        ch_infos: List[ChannelInfo] = []
        picks: List[int] = []
        for raw_idx, name in enumerate(labels):
            if name in self.task_config.exclude_channels:
                continue
            ch_type = _ch_type(name)
            override = self.task_config.channel_type_overrides.get(name)
            if override:
                ch_type = ChannelType(override)
            ch_infos.append(ChannelInfo(
                channel_id=f"ch_{len(ch_infos):03d}",
                index=len(ch_infos),
                name=name,
                standard_name=standardize_channel_name(name),
                type=ch_type,
                unit=self.task_config.signal_unit or "uV",
                sampling_rate=sfreq,
                status=ChannelStatus.UNKNOWN,
            ))
            picks.append(raw_idx)
        return ch_infos, picks

    def import_subject(
        self,
        cdt_path: Path,
        subject_id: Optional[str] = None,
        session_id: str = "ses-01",
        labels_map: Optional[Dict[int, Dict[str, str]]] = None,
        max_trials: Optional[int] = None,
        max_seconds: Optional[float] = None,
    ) -> ImportResult:
        from neuroatom.storage import paths as P
        from neuroatom.storage.metadata_store import AtomJSONLWriter
        from neuroatom.storage.signal_store import ShardManager

        cdt_path = Path(cdt_path)
        if subject_id is None:
            subject_id = cdt_path.stem  # "s1"
        sub_num = re.sub(r"\D", "", subject_id)

        dpa = _parse_dpa(Path(str(cdt_path) + ".dpa"))
        n_channels = dpa["n_channels"] or len(dpa["labels"])
        sfreq = dpa["sfreq"]
        labels = dpa["labels"]
        if len(labels) != n_channels:
            raise ValueError(
                f"{cdt_path.name}: {len(labels)} labels != {n_channels} channels"
            )

        if labels_map is None:
            xlsx = cdt_path.parent.parent / "group1.xlsx"
            all_labels = _parse_group_xlsx(xlsx) if xlsx.exists() else {}
            labels_map = all_labels.get(subject_id, {})

        data = _read_cdt(cdt_path, n_channels)  # (n_samples, n_channels) µV
        n_total = data.shape[0]
        trig_idx = next((i for i, l in enumerate(labels) if l.upper() == "TRIGGER"), None)
        if trig_idx is None:
            raise ValueError(f"{cdt_path.name}: no Trigger channel found")
        onsets = _find_trial_onsets(np.asarray(data[:, trig_idx]))
        trial_nums = sorted(onsets)
        if max_trials is not None:
            trial_nums = trial_nums[:max_trials]

        trial_seconds = float(
            self.task_config.data.get("custom", {}).get("trial_duration_seconds", 120.0)
        )
        window = max_seconds if max_seconds is not None else trial_seconds
        win_samples = int(window * sfreq)

        channel_infos, picks = self._build_channel_infos(labels, sfreq)
        channel_ids = [ci.channel_id for ci in channel_infos]

        dataset_id = self.task_config.dataset_id
        self.pool.ensure_dataset(dataset_id)
        self.pool.ensure_subject(dataset_id, subject_id)
        self.pool.ensure_session(dataset_id, subject_id, session_id, sfreq)

        max_shard_mb = self.pool.config.get("storage", {}).get("max_shard_size_mb", 200.0)
        compression = self.pool.config.get("storage", {}).get("compression", "gzip")

        stored_atoms: List[Atom] = []
        all_warnings: List[str] = []
        poor = _POOR_TRIALS.get(sub_num, set())

        for trial_num in trial_nums:
            onset = onsets[trial_num]
            end = min(onset + win_samples, n_total)
            seg = np.asarray(data[onset:end, :], dtype=np.float32)[:, picks].T  # (n_ch, n_samp)
            seg = np.ascontiguousarray(seg)
            n_samples = seg.shape[1]

            signal, storage_unit, orig_unit = convert_to_storage_unit(
                seg, source_unit=dpa["unit"], pool_config=self.pool.config,
            )

            run_id = f"trial_{trial_num:02d}"
            self.pool.ensure_run(dataset_id, subject_id, session_id, run_id)

            info = labels_map.get(trial_num, {})
            attended = {"L": "left", "R": "right"}.get(info.get("attended", ""), "unknown")

            annotations: List[Any] = [
                CategoricalAnnotation(
                    annotation_id=f"ann_ear_{run_id}",
                    name="attended_ear", value=attended,
                ),
                NumericAnnotation(
                    annotation_id=f"ann_trial_{run_id}",
                    name="trial_number", numeric_value=float(trial_num),
                ),
            ]
            if info.get("wav"):
                annotations.append(TextAnnotation(
                    annotation_id=f"ann_wav_{run_id}",
                    name="stimulus_file", text_value=info["wav"],
                ))
            if trial_num in poor:
                annotations.append(CategoricalAnnotation(
                    annotation_id=f"ann_excl_{run_id}",
                    name="excluded", value="true",
                ))

            atom_id = compute_atom_id(
                dataset_id=dataset_id, subject_id=subject_id,
                session_id=session_id, run_id=run_id, onset_sample=0,
            )
            atom = Atom(
                atom_id=atom_id, atom_type=AtomType.TRIAL,
                dataset_id=dataset_id, subject_id=subject_id,
                session_id=session_id, run_id=run_id, trial_index=trial_num,
                signal_ref=SignalRef(
                    file_path="", internal_path="",
                    shape=(len(channel_ids), n_samples),
                ),
                temporal=TemporalInfo(
                    onset_sample=0, onset_seconds=0.0,
                    duration_samples=n_samples, duration_seconds=n_samples / sfreq,
                ),
                channel_ids=channel_ids, n_channels=len(channel_ids),
                sampling_rate=sfreq, signal_unit=storage_unit, original_unit=orig_unit,
                annotations=annotations,
                custom_fields={"attended_ear": attended, "trial_number": trial_num},
            )

            all_warnings.extend(validate_signal(
                signal=signal, atom_id=atom.atom_id,
                config=self.pool.config.get("import", {}), signal_unit=storage_unit,
            ))

            with ShardManager(
                pool_root=self.pool.root, dataset_id=dataset_id,
                subject_id=subject_id, session_id=session_id, run_id=run_id,
                max_shard_size_mb=max_shard_mb, compression=compression,
            ) as shard_mgr:
                atom.signal_ref = shard_mgr.write_atom_signal(atom.atom_id, signal, None)

            jsonl_path = P.atoms_jsonl_path(
                self.pool.root, dataset_id, subject_id, session_id, run_id,
            )
            with AtomJSONLWriter(jsonl_path) as writer:
                writer.write_atom(atom)

            self.pool.register_run(RunMeta(
                run_id=run_id, session_id=session_id, subject_id=subject_id,
                dataset_id=dataset_id, run_index=trial_num,
                task_type=self.task_config.task_type, n_events=0, n_trials=1,
                paradigm_details={"attended_ear": attended},
            ))
            self._write_channels_json(dataset_id, subject_id, session_id, channel_infos)
            stored_atoms.append(atom)

        logger.info(
            "Imported BASE-AAD-USTC subject %s: %d trials", subject_id, len(stored_atoms),
        )
        run_meta = RunMeta(
            run_id=stored_atoms[0].run_id if stored_atoms else "trial_01",
            session_id=session_id, subject_id=subject_id, dataset_id=dataset_id,
            task_type=self.task_config.task_type, n_trials=len(stored_atoms),
        )
        return ImportResult(
            atoms=stored_atoms, run_meta=run_meta,
            channel_infos=channel_infos, warnings=all_warnings,
        )

    def import_dataset(
        self,
        root: Path,
        subjects: Optional[List[str]] = None,
        max_subjects: Optional[int] = None,
        max_trials: Optional[int] = None,
        max_seconds: Optional[float] = None,
    ) -> List[ImportResult]:
        root = Path(root)
        eeg_dir = root / "EEG" if (root / "EEG").is_dir() else root
        xlsx = (root / "group1.xlsx")
        if not xlsx.exists():
            xlsx = eeg_dir.parent / "group1.xlsx"
        all_labels = _parse_group_xlsx(xlsx) if xlsx.exists() else {}

        dataset_id = self.task_config.dataset_id
        cdt_files = sorted(
            eeg_dir.glob("*.cdt"),
            key=lambda p: int(re.sub(r"\D", "", p.stem) or 0),
        )
        if subjects:
            cdt_files = [f for f in cdt_files if f.stem in set(subjects)]
        if max_subjects:
            cdt_files = cdt_files[:max_subjects]

        self.pool.register_dataset(DatasetMeta(
            dataset_id=dataset_id, name=self.task_config.dataset_name,
            task_types=[self.task_config.task_type],
            n_subjects=len(cdt_files), original_format="curry",
        ))

        results: List[ImportResult] = []
        for cdt in cdt_files:
            self.pool.register_subject(SubjectMeta(
                subject_id=cdt.stem, dataset_id=dataset_id,
            ))
            try:
                results.append(self.import_subject(
                    cdt, subject_id=cdt.stem,
                    labels_map=all_labels.get(cdt.stem, {}),
                    max_trials=max_trials, max_seconds=max_seconds,
                ))
            except Exception as e:
                logger.error("Failed to import %s: %s", cdt.name, e)

        try:
            self.pool.assess_quality(dataset_id)
        except Exception:
            pass

        logger.info("BASE-AAD-USTC import complete: %d subjects", len(results))
        return results


# Auto-register
register_importer("base_aad_ustc", BASEAADUSTCImporter)
